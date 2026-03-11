import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook # type: ignore
from bs4 import BeautifulSoup

def get_confirmation():
    choice = input("Do you want to create yesterday's report? (Press 'N' or 'n' for No, any other key for Yes): ").strip()
    return choice.lower() != 'n'

current_month = pd.Timestamp('today').strftime('%b')
current_month_num = pd.Timestamp('today').strftime('%m')

if get_confirmation():
    today = datetime.now() - timedelta(days=1)
    today = today.strftime("%d-%b-%y")  # Example: 10-Mar-25
else:
    date = int(input("Enter how many days ago from today's date you want to get: "))
    date = datetime.now() - timedelta(days=date)
    today = date.strftime("%d-%b-%y")

file_path = f"F:/1. Work/1. Daily/Production follow up/{current_month_num}. {current_month}/{today}.xlsx"
production_followup_wb = load_workbook(file_path, data_only=True)
plan_pcs = production_followup_wb.active['B17'].value
accu_qc_pass = production_followup_wb.active['C17'].value
plan_bal = plan_pcs - accu_qc_pass
qc_pass = production_followup_wb.active['B5'].value
efficiency = production_followup_wb.active['G5'].value
completed_days = production_followup_wb.active['O5'].value - production_followup_wb.active['P5'].value
req_per_day = int((qc_pass + plan_bal) / (completed_days + 1))
operator_present = production_followup_wb.active['K5'].value

# cutting data
cutting_file = 'DBL GROUP.html'
with open(cutting_file, "r", encoding="utf-8") as file:
    html_file = file.read()
html_content = BeautifulSoup(html_file, "html.parser")
rows = []
for row in html_content.find_all('tr'):
    row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
    rows.append(row_data)
cutting_pcs = ''
for row in rows:
    if len(row) == 8 and row[1] == 'JFL SubTotal':
        cutting_pcs = row[2]

# print Embroidery
emb_pcs = 0
print_pcs = 0
print_emb = 0
print_emb_wb = load_workbook("JFL-PRINT & EMB BALANCE STATUS.xlsx", data_only=True)
print_emb_ws = print_emb_wb.active
for i in range(9, 1000):
    if print_emb_ws[f'K{i}'].value != None:
        if print_emb_ws[f'E{i}'].value == "PRINT":
            print_pcs += int(print_emb_ws[f'K{i}'].value)
        elif print_emb_ws[f'E{i}'].value == "EMB":
            emb_pcs += int(print_emb_ws[f'K{i}'].value)
        elif print_emb_ws[f'E{i}'].value == "PRINT+EMB":
            print_emb += int(print_emb_ws[f'K{i}'].value)

# Fabric Status
master_file = load_workbook("F:/1. Work/1. Daily/Master File JFL/Master File-JFL .xlsx", data_only=True)
master_ws = master_file["Fabric Main"]
buyer_fabric = {}
total_fabric_rec = 0
for i in range(3, 1000):
    if master_ws[f'M{i}'].value != None and master_ws[f'M{i}'].value > 0:
        total_fabric_rec += master_ws[f'M{i}'].value
        if master_ws[f'E{i}'].value not in buyer_fabric:
            buyer_fabric[master_ws[f'E{i}'].value] = 0
        buyer_fabric[master_ws[f'E{i}'].value] += master_ws[f'M{i}'].value
print(buyer_fabric)
buyer_fabric_text = ""
for buyer, qty in buyer_fabric.items():
    buyer_fabric_text += f"{buyer:<20} {qty:>10} Kg\n"


result = f"""
Production & Fabric Challan Summary of JFL on {today}.

{current_month} plan quantity= {plan_pcs} pcs
Till now production quantity= {accu_qc_pass} pcs
Production balance= {plan_bal} pcs
Required per day= {req_per_day} pcs
Cutting= {cutting_pcs} Pcs
Output= {qc_pass} Pcs
Efficiency= {efficiency * 100} %

Total Fabric challan = {total_fabric_rec} Kg 
{buyer_fabric_text}
{today} Operator Status:
Existing: 608
Present: {operator_present}

Print & Emb. Rcvd Status on {today}
Print= {print_pcs + print_emb} Pcs
Emb.= {emb_pcs + print_emb} Pcs
"""
with open(f"{today}.txt", "w", encoding="utf-8") as f:
    f.write(result)
